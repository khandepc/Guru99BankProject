import openpyxl

def read_data():
    list=[]
    path = "C:\All_Desktop_Data_2020\Gure99BankProject\TestDataManagerLogin.xlsx"
    workbook=openpyxl.load_workbook(path)
    sheet=workbook["Sheet1"]

    rows=sheet.max_row
    cols=sheet.max_column

    for r in range(2,rows+1):
        # for c in range(1,cols+1):
        #     print(sheet.cell(r,c).value)
        userid=sheet.cell(r,1).value
        password=sheet.cell(r,2).value
        tuple=(userid,password)
        list.append(tuple)
    return list
